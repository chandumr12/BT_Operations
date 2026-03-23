from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
import uuid
import re
import io
import os
import base64
import json
import asyncio
import traceback
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
from datetime import datetime, timezone
import firebase_admin
from firebase_admin import credentials, auth as firebase_auth, firestore
from openpyxl import Workbook, load_workbook

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# -----------------------------
# Firebase Initialization
# -----------------------------
if not firebase_admin._apps:
    try:
        firebase_key_base64 = os.getenv("FIREBASE_KEY_BASE64")

        if firebase_key_base64:
            # 🔥 Production (Cloud Run)
            decoded_key = base64.b64decode(firebase_key_base64)
            cred = credentials.Certificate(
                json.loads(decoded_key.decode("utf-8"))
            )
            print("✅ Firebase initialized using BASE64 (Production)")

        else:
            # 🔥 Local Development
            cred_path = os.getenv(
                "FIREBASE_CREDENTIALS_PATH",
                str(ROOT_DIR / "firebase-admin.json")
            )

            if not os.path.exists(cred_path):
                raise Exception(f"Firebase credentials file not found at {cred_path}")

            cred = credentials.Certificate(cred_path)
            print("✅ Firebase initialized using LOCAL JSON")

        firebase_admin.initialize_app(cred)
        db = firestore.client()
        print("✅ Firestore connected successfully")

    except Exception as e:
        print("❌ Firebase initialization failed:")
        print(str(e))
        db = None

# Firestore client
try:
    db = firestore.client()
except:
    db = None
    logging.warning("Firestore client not initialized")

# -------------------------------------------------------
# Finance Firestore client (bt-finance project)
# -------------------------------------------------------
finance_db = None
try:
    FINANCE_KEY_BASE64 = os.getenv("FINANCE_FIREBASE_KEY_BASE64")
    FINANCE_CRED_PATH = os.getenv("FINANCE_FIREBASE_CREDENTIALS_PATH",
                                   str(Path(__file__).parent / "firebase-finance.json"))
    if FINANCE_KEY_BASE64:
        finance_decoded = base64.b64decode(FINANCE_KEY_BASE64)
        finance_cred = credentials.Certificate(json.loads(finance_decoded.decode("utf-8")))
        print("✅ Finance Firebase initialized using BASE64")
    elif os.path.exists(FINANCE_CRED_PATH):
        finance_cred = credentials.Certificate(FINANCE_CRED_PATH)
        print("✅ Finance Firebase initialized using LOCAL JSON")
    else:
        finance_cred = None
        print("⚠️  Finance Firebase credentials not found — payout routes will return empty")
    if finance_cred:
        if 'finance' not in [a.name for a in firebase_admin._apps.values()]:
            finance_app = firebase_admin.initialize_app(finance_cred, name='finance')
        else:
            finance_app = firebase_admin.get_app('finance')
        finance_db = firestore.client(app=finance_app)
        print("✅ Finance Firestore connected")
except Exception as e:
    finance_db = None
    print(f"⚠️  Finance Firestore not connected: {e}")

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# File Storage - local filesystem
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

APP_NAME = "bengaluru-trekkers"
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

def put_object(path: str, data: bytes, content_type: str) -> dict:
    file_path = UPLOAD_DIR / path
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_bytes(data)
    return {"path": path}

def get_object(path: str):
    file_path = UPLOAD_DIR / path
    if not file_path.exists():
        raise Exception("File not found")
    content = file_path.read_bytes()
    import mimetypes
    content_type = mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
    return content, content_type

def delete_object(path: str):
    file_path = UPLOAD_DIR / path
    if file_path.exists():
        file_path.unlink()

# Constants
HARDCODED_ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'admin@bengalurutrekkers.com')
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")

# Models
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    displayName: str

class UserResponse(BaseModel):
    uid: str
    email: str
    displayName: str
    role: str
    status: str
    createdAt: str

class TrekCreate(BaseModel):
    name: str
    region: str
    distanceFromBengaluru: str
    trekDistance: str
    altitude: str
    difficultyLevel: str
    bestTimeToVisit: str
    meetingPoint: str
    reportingTime: str
    requiredPermissions: str = ""
    vendorNotes: str = ""
    internalNotes: str = ""
    trekType: str
    category: str

class TrekUpdate(BaseModel):
    name: Optional[str] = None
    region: Optional[str] = None
    distanceFromBengaluru: Optional[str] = None
    trekDistance: Optional[str] = None
    altitude: Optional[str] = None
    difficultyLevel: Optional[str] = None
    bestTimeToVisit: Optional[str] = None
    meetingPoint: Optional[str] = None
    reportingTime: Optional[str] = None
    requiredPermissions: Optional[str] = None
    vendorNotes: Optional[str] = None
    internalNotes: Optional[str] = None
    trekType: Optional[str] = None
    category: Optional[str] = None
    archived: Optional[bool] = None

class BatchCreate(BaseModel):
    batchCode: str
    trekId: str
    startDate: str
    endDate: str
    maxCapacity: int
    currentRegistrations: int = 0
    status: str = "Open"
    assignedTrekLead: Optional[str] = None
    assignedSupportLead: Optional[str] = None
    assignedCoordinator: Optional[str] = None
    assignedLeads: List[dict] = []  # [{userId, displayName, isSuperLead}]
    transportVendor: Optional[str] = ""
    stayVendor: Optional[str] = ""
    internalNotes: Optional[str] = ""

class LeadCreate(BaseModel):
    name: str
    phone: str
    age: int
    gender: str
    active: bool = True
    hiredDate: str
    languages: List[str] = []
    specialSkills: List[str] = []
    email: Optional[str] = None
    password: Optional[str] = None
    role: str = "Trek Lead"
    # Finance / KYC fields — all Optional so existing leads are unaffected
    leadType: Optional[str] = None
    notes: Optional[str] = None
    bankName: Optional[str] = None
    accountName: Optional[str] = None
    accountNumber: Optional[str] = None
    ifsc: Optional[str] = None
    panNumber: Optional[str] = None

class LeadSelfRegister(BaseModel):
    name: str
    phone: str
    age: int
    gender: str
    hiredDate: str
    languages: List[str] = []
    specialSkills: List[str] = []
    email: str
    password: str

class ChecklistCreate(BaseModel):
    batchId: str
    type: str  # pre, during, post
    items: List[dict]

class TicketCreate(BaseModel):
    title: str
    description: str = ""
    priority: str = "Medium"  # Low, Medium, High, Urgent
    status: str = "Backlog"  # Backlog, To Do, In Progress, In Review, Done, Blocked
    category: str = "Operations"  # Operations, Sales, Content, Development, Trek Planning
    assignees: List[str] = []  # Changed to support multiple assignees
    dueDate: Optional[str] = None
    estimatedHours: Optional[int] = None
    labels: List[str] = []
    attachments: List[dict] = []  # {name, url, uploadedBy, uploadedAt}

class TicketUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    category: Optional[str] = None
    assignees: Optional[List[str]] = None
    dueDate: Optional[str] = None
    estimatedHours: Optional[int] = None
    labels: Optional[List[str]] = None
    attachments: Optional[List[dict]] = None

# Firestore helper — wraps any blocking Firestore call with a timeout.
# Returns `default` on timeout/error so endpoints degrade gracefully locally.
async def fs_run(fn, timeout=4.0, default=None):
    try:
        loop = asyncio.get_running_loop()
        return await asyncio.wait_for(loop.run_in_executor(None, fn), timeout=timeout)
    except asyncio.TimeoutError:
        print(f"[FIRESTORE] Query timed out after {timeout}s")
        return default
    except Exception as e:
        print(f"[FIRESTORE] Error: {type(e).__name__}: {e}")
        return default

# Email helper — sends batch assignment email via Gmail SMTP in a background thread.
async def send_assignment_email(to_email: str, lead_name: str, role_label: str,
                                batch_code: str, trek_name: str,
                                start_date: str, end_date: str, assigner: str):
    if not SMTP_USER or not SMTP_PASSWORD:
        print("[EMAIL] SMTP not configured — skipping email")
        return
    subject = f"BT Ops: You've been assigned to Batch {batch_code}"
    body = (
        f"Hi {lead_name},\n\n"
        f"You have been assigned to a new trek batch on BT Ops.\n\n"
        f"  Trek      : {trek_name}\n"
        f"  Batch     : {batch_code}\n"
        f"  Your Role : {role_label}\n"
        f"  Dates     : {start_date} → {end_date}\n"
        f"  Assigned by: {assigner}\n\n"
        f"Please log in to BT Ops to view full batch details.\n\n"
        f"— Bengaluru Trekkers Operations\n"
    )
    msg = MIMEMultipart()
    msg["From"] = SMTP_USER
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    def _send():
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(SMTP_USER, to_email, msg.as_string())

    try:
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, _send)
        print(f"[EMAIL] Sent assignment email to {to_email}")
    except Exception as e:
        print(f"[EMAIL] Failed to send to {to_email}: {e}")
        # Non-fatal — batch save must not fail because of email

# Authentication Dependency
async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        loop = asyncio.get_running_loop()
        decoded_token = await asyncio.wait_for(
            loop.run_in_executor(
                None, lambda: firebase_auth.verify_id_token(credentials.credentials)
            ),
            timeout=8.0
        )
        uid = decoded_token['uid']
        email = decoded_token.get('email', '')
        now = datetime.now(timezone.utc).isoformat()
        is_admin = email == HARDCODED_ADMIN_EMAIL

        if db:
            try:
                user_ref = db.collection('users').document(uid)
                user_doc = await asyncio.wait_for(
                    loop.run_in_executor(None, user_ref.get),
                    timeout=3.0
                )
                if user_doc.exists:
                    user_data = user_doc.to_dict()
                    user_data['uid'] = uid
                    user_data.setdefault('displayName', email.split('@')[0])
                    user_data.setdefault('createdAt', now)
                    return user_data
                else:
                    doc = {
                        'email': email,
                        'displayName': decoded_token.get('name', email.split('@')[0]),
                        'role': 'Super Admin' if is_admin else 'Coordinator',
                        'status': 'approved' if is_admin else 'pending',
                        'createdAt': now,
                    }
                    await loop.run_in_executor(None, lambda: user_ref.set(doc))
                    doc['uid'] = uid
                    return doc
            except asyncio.TimeoutError:
                print(f"[AUTH] Firestore timed out for {uid} — using token fallback")
            except Exception as db_err:
                print(f"[AUTH] Firestore error: {type(db_err).__name__}: {db_err}")

        # Fallback: derive profile from token
        return {
            'uid': uid,
            'email': email,
            'displayName': decoded_token.get('name', email.split('@')[0]),
            'role': 'Super Admin' if is_admin else 'user',
            'status': 'approved',
            'createdAt': now,
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[AUTH ERROR] {type(e).__name__}: {e}")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"Invalid authentication credentials: {str(e)}"
        )

# Admin-only dependency
async def require_admin(user: dict = Depends(get_current_user)):
    if user.get('role') != 'Super Admin':
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return user

# Routes
@api_router.get("/")
async def root():
    return {"message": "Bengaluru Trekkers Operations Management System API"}

@api_router.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "firebase_initialized": firebase_admin._apps != {},
        "firestore_connected": db is not None
    }

# Auth Routes
@api_router.post("/auth/register")
async def register_user(user_data: UserCreate):
    try:
        # Create user in Firebase Auth
        user_record = firebase_auth.create_user(
            email=user_data.email,
            password=user_data.password,
            display_name=user_data.displayName
        )
        
        # Determine role and status
        is_admin = user_data.email == HARDCODED_ADMIN_EMAIL
        
        # Create user document in Firestore
        user_doc = {
            'email': user_data.email,
            'displayName': user_data.displayName,
            'role': 'Super Admin' if is_admin else 'Coordinator',
            'status': 'approved' if is_admin else 'pending',
            'createdAt': datetime.now(timezone.utc).isoformat()
        }
        
        if db:
            db.collection('users').document(user_record.uid).set(user_doc)
        
        return {
            'uid': user_record.uid,
            'message': 'Admin account created' if is_admin else 'Registration pending admin approval'
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_current_user_info(user: dict = Depends(get_current_user)):
    return user

# Basic user list for all authenticated users (for assignee display)
@api_router.get("/users/basic")
async def list_users_basic(user: dict = Depends(get_current_user)):
    if not db:
        return []
    users_ref = db.collection('users')
    users = users_ref.stream()
    return [
        {'uid': doc.id, 'displayName': doc.to_dict().get('displayName', ''), 'role': doc.to_dict().get('role', ''), 'status': doc.to_dict().get('status', '')}
        for doc in users
    ]

# User Management Routes (Admin only)
@api_router.get("/users")
async def list_users(user: dict = Depends(require_admin)):
    if not db:
        return []
    
    users_ref = db.collection('users')
    users = users_ref.stream()
    
    user_list = []
    for user_doc in users:
        user_data = user_doc.to_dict()
        user_data['uid'] = user_doc.id
        user_list.append(user_data)
    
    return user_list

@api_router.patch("/users/{uid}/approve")
async def approve_user(uid: str, admin: dict = Depends(require_admin)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    user_ref = db.collection('users').document(uid)
    user_ref.update({'status': 'approved'})
    
    return {'message': 'User approved successfully'}

@api_router.patch("/users/{uid}/role")
async def update_user_role(uid: str, role: str, admin: dict = Depends(require_admin)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    user_ref = db.collection('users').document(uid)
    user_ref.update({'role': role})
    
    return {'message': 'User role updated successfully'}

@api_router.delete("/users/{uid}")
async def delete_user(uid: str, admin: dict = Depends(require_admin)):
    try:
        firebase_auth.delete_user(uid)
        if db:
            db.collection('users').document(uid).delete()
        return {'message': 'User deleted successfully'}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Trek Routes
@api_router.post("/treks")
async def create_trek(trek: TrekCreate, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    trek_data = trek.model_dump()
    trek_data['archived'] = False
    trek_data['createdAt'] = datetime.now(timezone.utc).isoformat()
    trek_data['createdBy'] = user['uid']
    
    doc_ref = db.collection('treks').document()
    doc_ref.set(trek_data)
    
    trek_data['id'] = doc_ref.id
    return trek_data

@api_router.get("/treks")
async def list_treks(user: dict = Depends(get_current_user), include_archived: bool = False):
    if not db:
        return []
    
    treks_ref = db.collection('treks')

    if not include_archived:
        treks_ref = treks_ref.where('archived', '==', False)

    treks = await fs_run(lambda: list(treks_ref.stream()), default=[])

    trek_list = []
    for trek_doc in treks:
        trek_data = trek_doc.to_dict()
        trek_data['id'] = trek_doc.id
        trek_list.append(trek_data)
    
    return trek_list

@api_router.get("/treks/{trek_id}")
async def get_trek(trek_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    trek_doc = db.collection('treks').document(trek_id).get()
    
    if not trek_doc.exists:
        raise HTTPException(status_code=404, detail="Trek not found")
    
    trek_data = trek_doc.to_dict()
    trek_data['id'] = trek_doc.id
    return trek_data

@api_router.patch("/treks/{trek_id}")
async def update_trek(trek_id: str, trek_update: TrekUpdate, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    trek_ref = db.collection('treks').document(trek_id)
    
    update_data = {k: v for k, v in trek_update.model_dump().items() if v is not None}
    update_data['updatedAt'] = datetime.now(timezone.utc).isoformat()
    
    trek_ref.update(update_data)
    
    return {'message': 'Trek updated successfully', 'id': trek_id}

# Batch Routes
async def create_notifications_for_batch(batch_id, batch_code, trek_name, assigned_leads,
                                         assigner_name, start_date="", end_date=""):
    """Create in-app notifications and send email for each newly assigned lead."""
    if not db or not assigned_leads:
        return
    for lead in assigned_leads:
        role_label = "Trek Lead"
        # In-app notification
        db.collection('notifications').add({
            'userId': lead['userId'],
            'type': 'batch_assignment',
            'title': f'Assigned to Batch {batch_code}',
            'message': f'You have been assigned as {role_label} for batch {batch_code} ({trek_name})',
            'batchId': batch_id,
            'read': False,
            'createdAt': datetime.now(timezone.utc).isoformat(),
            'createdBy': assigner_name,
        })
        # Email notification
        user_doc = await fs_run(
            lambda uid=lead['userId']: db.collection('users').document(uid).get()
        )
        if user_doc and user_doc.exists:
            email = user_doc.to_dict().get('email', '')
            if email:
                await send_assignment_email(
                    to_email=email,
                    lead_name=lead.get('displayName', ''),
                    role_label=role_label,
                    batch_code=batch_code,
                    trek_name=trek_name,
                    start_date=start_date,
                    end_date=end_date,
                    assigner=assigner_name,
                )

@api_router.post("/batches")
async def create_batch(batch: BatchCreate, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    existing = db.collection('batches').where('batchCode', '==', batch.batchCode).limit(1).stream()
    if list(existing):
        raise HTTPException(status_code=400, detail="Batch code already exists")
    
    batch_data = batch.model_dump()
    batch_data['createdAt'] = datetime.now(timezone.utc).isoformat()
    batch_data['createdBy'] = user['uid']
    
    doc_ref = db.collection('batches').document()
    doc_ref.set(batch_data)
    batch_data['id'] = doc_ref.id

    # Get trek name for notification
    trek_name = ""
    if batch.trekId:
        trek_doc = db.collection('treks').document(batch.trekId).get()
        if trek_doc.exists:
            trek_name = trek_doc.to_dict().get('name', '')

    if batch_data.get('assignedLeads'):
        await create_notifications_for_batch(
            doc_ref.id, batch.batchCode, trek_name,
            batch_data['assignedLeads'], user.get('displayName', 'Admin'),
            start_date=batch.startDate,
            end_date=batch.endDate,
        )
    
    return batch_data

@api_router.get("/batches/my")
async def get_my_batches(user: dict = Depends(get_current_user)):
    """Get batches assigned to the current user (for Trek Leads)."""
    if not db:
        return []
    all_batches = await fs_run(lambda: list(db.collection('batches').stream()), default=[])
    result = []
    for batch_doc in all_batches:
        batch_data = batch_doc.to_dict()
        batch_data['id'] = batch_doc.id
        assigned_leads = batch_data.get('assignedLeads', [])
        for lead in assigned_leads:
            if lead.get('userId') == user['uid']:
                batch_data['myRole'] = 'Trek Lead'
                result.append(batch_data)
                break
    return result

@api_router.get("/batches")
async def list_batches(user: dict = Depends(get_current_user), trek_id: Optional[str] = None):
    if not db:
        return []
    
    batches_ref = db.collection('batches')
    if trek_id:
        batches_ref = batches_ref.where('trekId', '==', trek_id)

    all_batches = await fs_run(lambda: list(batches_ref.stream()), default=[])
    user_role = user.get('role', '')
    
    batch_list = []
    for batch_doc in all_batches:
        batch_data = batch_doc.to_dict()
        batch_data['id'] = batch_doc.id
        
        # Trek Leads only see their assigned batches
        if user_role == 'Trek Lead':
            assigned_leads = batch_data.get('assignedLeads', [])
            is_assigned = any(l.get('userId') == user['uid'] for l in assigned_leads)
            # Also check legacy fields
            if not is_assigned and batch_data.get('assignedTrekLead') != user['uid'] and batch_data.get('assignedSupportLead') != user['uid']:
                continue
        
        batch_list.append(batch_data)
    
    return batch_list

@api_router.get("/batches/{batch_id}")
async def get_batch(batch_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    batch_doc = db.collection('batches').document(batch_id).get()
    if not batch_doc.exists:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    batch_data = batch_doc.to_dict()
    batch_data['id'] = batch_doc.id
    return batch_data

@api_router.patch("/batches/{batch_id}")
async def update_batch(batch_id: str, batch_data: dict, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    batch_ref = db.collection('batches').document(batch_id)
    old_batch_doc = batch_ref.get()
    old_batch = old_batch_doc.to_dict() if old_batch_doc.exists else {}
    
    batch_data['updatedAt'] = datetime.now(timezone.utc).isoformat()
    batch_ref.update(batch_data)
    
    # Notify newly assigned leads
    if 'assignedLeads' in batch_data:
        old_lead_ids = {l['userId'] for l in old_batch.get('assignedLeads', [])}
        new_leads = [l for l in batch_data['assignedLeads'] if l['userId'] not in old_lead_ids]
        if new_leads:
            trek_name = ""
            trek_id = batch_data.get('trekId', old_batch.get('trekId', ''))
            if trek_id:
                trek_doc = db.collection('treks').document(trek_id).get()
                if trek_doc.exists:
                    trek_name = trek_doc.to_dict().get('name', '')
            batch_code = batch_data.get('batchCode', old_batch.get('batchCode', ''))
            await create_notifications_for_batch(
                batch_id, batch_code, trek_name,
                new_leads, user.get('displayName', 'Admin'),
                start_date=batch_data.get('startDate', old_batch.get('startDate', '')),
                end_date=batch_data.get('endDate', old_batch.get('endDate', '')),
            )
    
    return {'message': 'Batch updated successfully', 'id': batch_id}

@api_router.delete("/batches/{batch_id}")
async def delete_batch(batch_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    allowed_roles = ['Super Admin', 'Operations Manager']
    if user.get('role') not in allowed_roles:
        raise HTTPException(status_code=403, detail="Only admins can delete batches")
    batch_ref = db.collection('batches').document(batch_id)
    if not batch_ref.get().exists:
        raise HTTPException(status_code=404, detail="Batch not found")
    batch_ref.delete()
    return {'message': 'Batch deleted successfully'}

# Team Member Routes
@api_router.get("/team")
async def list_team(user: dict = Depends(get_current_user)):
    if not db: return []
    docs = db.collection('team_members').stream()
    return [{**d.to_dict(), 'id': d.id} for d in docs]

@api_router.post("/team")
async def create_team_member(data: dict, user: dict = Depends(get_current_user)):
    if user.get('role') not in ['Super Admin', 'Operations Manager']:
        raise HTTPException(status_code=403, detail="Admins only")
    data['createdAt'] = datetime.now(timezone.utc).isoformat()
    ref = db.collection('team_members').add(data)
    return {'id': ref[1].id, **data}

@api_router.patch("/team/{member_id}")
async def update_team_member(member_id: str, data: dict, user: dict = Depends(get_current_user)):
    if user.get('role') not in ['Super Admin', 'Operations Manager']:
        raise HTTPException(status_code=403, detail="Admins only")
    data['updatedAt'] = datetime.now(timezone.utc).isoformat()
    db.collection('team_members').document(member_id).update(data)
    return {'message': 'Updated', 'id': member_id}

@api_router.delete("/team/{member_id}")
async def delete_team_member(member_id: str, user: dict = Depends(get_current_user)):
    if user.get('role') not in ['Super Admin', 'Operations Manager']:
        raise HTTPException(status_code=403, detail="Admins only")
    db.collection('team_members').document(member_id).delete()
    return {'message': 'Deleted'}

# Notification Routes
@api_router.get("/notifications")
async def get_notifications(user: dict = Depends(get_current_user)):
    if not db:
        return []
    notifs = db.collection('notifications').where('userId', '==', user['uid']).stream()
    result = []
    for doc in notifs:
        data = doc.to_dict()
        data['id'] = doc.id
        result.append(data)
    result.sort(key=lambda x: x.get('createdAt', ''), reverse=True)
    return result

@api_router.get("/notifications/unread-count")
async def get_unread_count(user: dict = Depends(get_current_user)):
    if not db:
        return {'count': 0}
    notifs = db.collection('notifications').where('userId', '==', user['uid']).where('read', '==', False).stream()
    return {'count': len(list(notifs))}

@api_router.patch("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    db.collection('notifications').document(notification_id).update({'read': True})
    return {'message': 'Notification marked as read'}

@api_router.patch("/notifications/read-all")
async def mark_all_notifications_read(user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    notifs = db.collection('notifications').where('userId', '==', user['uid']).where('read', '==', False).stream()
    for doc in notifs:
        db.collection('notifications').document(doc.id).update({'read': True})
    return {'message': 'All notifications marked as read'}

# Participant Routes
PARTICIPANT_TEMPLATE_HEADERS = [
    'SL No', 'Full Name', 'Contact No', 'Age', 'Gender', 'Pickup Point',
    'Total Price incl GST', 'Amount Paid', 'Balance Amount', 'Receipt Mode',
    'Receipt Date', 'If Cancelled Reason', 'Booked By', 'Remarks'
]

@api_router.get("/batches/{batch_id}/participants/template")
async def download_participant_template(batch_id: str, user: dict = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws.append(PARTICIPANT_TEMPLATE_HEADERS)
    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    # Set column widths
    widths = [8, 25, 15, 6, 10, 20, 18, 14, 14, 14, 14, 20, 15, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=participant_template_{batch_id}.xlsx"}
    )

@api_router.post("/batches/{batch_id}/participants/import")
async def import_participants(batch_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    batch_doc = db.collection('batches').document(batch_id).get()
    if not batch_doc.exists:
        raise HTTPException(status_code=404, detail="Batch not found")

    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in the file")

    imported = []
    for row in rows:
        if not row or not row[1]:  # Skip empty rows (check Full Name column)
            continue
        def safe_str(v): return str(v).strip() if v is not None else ''
        def safe_float(v):
            if v is None or str(v).strip() == '': return 0.0
            try: return float(str(v).replace(',', ''))
            except: return 0.0

        participant = {
            'slNo': safe_str(row[0]) if len(row) > 0 else '',
            'fullName': safe_str(row[1]) if len(row) > 1 else '',
            'contactNo': safe_str(row[2]) if len(row) > 2 else '',
            'age': safe_str(row[3]) if len(row) > 3 else '',
            'gender': safe_str(row[4]) if len(row) > 4 else '',
            'pickupPoint': safe_str(row[5]) if len(row) > 5 else '',
            'totalPrice': safe_float(row[6]) if len(row) > 6 else 0,
            'amountPaid': safe_float(row[7]) if len(row) > 7 else 0,
            'balanceAmount': safe_float(row[8]) if len(row) > 8 else 0,
            'receiptMode': safe_str(row[9]) if len(row) > 9 else '',
            'receiptDate': safe_str(row[10]) if len(row) > 10 else '',
            'cancelledReason': safe_str(row[11]) if len(row) > 11 else '',
            'bookedBy': safe_str(row[12]) if len(row) > 12 else '',
            'remarks': safe_str(row[13]) if len(row) > 13 else '',
            'batchId': batch_id,
            'boarded': False,
            'amountCollected': 0.0,
            'status': 'Confirmed',
            'createdAt': datetime.now(timezone.utc).isoformat(),
            'createdBy': user.get('displayName', 'Unknown'),
        }
        doc_ref = db.collection('participants').document()
        doc_ref.set(participant)
        participant['id'] = doc_ref.id
        imported.append(participant)

    return {'message': f'{len(imported)} participants imported', 'count': len(imported), 'participants': imported}

@api_router.get("/batches/{batch_id}/participants")
async def list_participants(batch_id: str, user: dict = Depends(get_current_user)):
    if not db:
        return []
    docs = db.collection('participants').where('batchId', '==', batch_id).stream()
    result = []
    for doc in docs:
        data = doc.to_dict()
        data['id'] = doc.id
        result.append(data)
    result.sort(key=lambda x: x.get('slNo', ''))
    return result

@api_router.post("/batches/{batch_id}/participants")
async def add_participant(batch_id: str, participant: dict, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    batch_doc = db.collection('batches').document(batch_id).get()
    if not batch_doc.exists:
        raise HTTPException(status_code=404, detail="Batch not found")
    participant['batchId'] = batch_id
    participant['boarded'] = participant.get('boarded', False)
    participant['amountCollected'] = participant.get('amountCollected', 0.0)
    participant['status'] = participant.get('status', 'Confirmed')
    participant['createdAt'] = datetime.now(timezone.utc).isoformat()
    participant['createdBy'] = user.get('displayName', 'Unknown')
    doc_ref = db.collection('participants').document()
    doc_ref.set(participant)
    participant['id'] = doc_ref.id
    return participant

@api_router.patch("/batches/{batch_id}/participants/{participant_id}")
async def update_participant(batch_id: str, participant_id: str, data: dict, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    ref = db.collection('participants').document(participant_id)
    doc = ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Participant not found")
    data['updatedAt'] = datetime.now(timezone.utc).isoformat()
    data['updatedBy'] = user.get('displayName', 'Unknown')
    ref.update(data)
    return {'message': 'Participant updated', 'id': participant_id}

@api_router.delete("/batches/{batch_id}/participants/{participant_id}")
async def delete_participant(batch_id: str, participant_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    db.collection('participants').document(participant_id).delete()
    return {'message': 'Participant removed'}

# Batch Expense Routes
EXPENSE_FIELDS = [
    'amountCollected', 'paidToDriver', 'lunchPacking', 'parkingCharges',
    'jeepCharges', 'refundToCustomer', 'tickets', 'localGuide',
    'leadsLunchExpenses', 'otherExpenses', 'otherExpensesRemarks'
]

def calc_expense_totals(data):
    fixed_spent = sum(data.get(f, 0) or 0 for f in EXPENSE_FIELDS[1:10])
    additional = data.get('additionalExpenses', [])
    additional_spent = sum(float(item.get('amount', 0) or 0) for item in additional)
    data['totalSpent'] = fixed_spent + additional_spent
    data['remaining'] = (data.get('amountCollected', 0) or 0) - data['totalSpent']
    return data

@api_router.get("/batches/{batch_id}/expenses")
async def list_batch_expenses(batch_id: str, user: dict = Depends(get_current_user)):
    if not db:
        return []
    docs = db.collection('batch_expenses').where('batchId', '==', batch_id).stream()
    result = []
    for doc in docs:
        data = doc.to_dict()
        data['id'] = doc.id
        calc_expense_totals(data)
        result.append(data)
    return result

@api_router.get("/batches/{batch_id}/expenses/my")
async def get_my_expense(batch_id: str, user: dict = Depends(get_current_user)):
    if not db:
        return None
    docs = list(db.collection('batch_expenses').where('batchId', '==', batch_id).where('userId', '==', user['uid']).limit(1).stream())
    if not docs:
        return None
    data = docs[0].to_dict()
    data['id'] = docs[0].id
    calc_expense_totals(data)
    return data

@api_router.post("/batches/{batch_id}/expenses")
async def create_or_update_expense(batch_id: str, expense_data: dict, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    batch_doc = db.collection('batches').document(batch_id).get()
    if not batch_doc.exists:
        raise HTTPException(status_code=404, detail="Batch not found")

    existing = list(db.collection('batch_expenses').where('batchId', '==', batch_id).where('userId', '==', user['uid']).limit(1).stream())

    clean = {}
    for f in EXPENSE_FIELDS:
        if f == 'otherExpensesRemarks':
            clean[f] = str(expense_data.get(f, ''))
        else:
            val = expense_data.get(f, 0)
            try:
                clean[f] = float(str(val).replace(',', '')) if val else 0.0
            except:
                clean[f] = 0.0

    # Handle additional custom expenses
    additional = expense_data.get('additionalExpenses', [])
    clean_additional = []
    for item in additional:
        if item.get('reason') or item.get('amount'):
            try:
                amt = float(str(item.get('amount', 0)).replace(',', '')) if item.get('amount') else 0.0
            except:
                amt = 0.0
            clean_additional.append({'reason': str(item.get('reason', '')), 'amount': amt})
    clean['additionalExpenses'] = clean_additional

    clean['batchId'] = batch_id
    clean['userId'] = user['uid']
    clean['leadName'] = user.get('displayName', 'Unknown')
    clean['updatedAt'] = datetime.now(timezone.utc).isoformat()

    if existing:
        doc_id = existing[0].id
        db.collection('batch_expenses').document(doc_id).update(clean)
        clean['id'] = doc_id
    else:
        clean['createdAt'] = datetime.now(timezone.utc).isoformat()
        ref = db.collection('batch_expenses').document()
        ref.set(clean)
        clean['id'] = ref.id

    calc_expense_totals(clean)
    return clean

@api_router.delete("/batches/{batch_id}/expenses/{expense_id}")
async def delete_expense(batch_id: str, expense_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    db.collection('batch_expenses').document(expense_id).delete()
    return {'message': 'Expense deleted'}

# Batch Document Routes
@api_router.get("/batches/{batch_id}/documents")
async def list_batch_documents(batch_id: str, user: dict = Depends(get_current_user)):
    if not db:
        return []
    docs = db.collection('batch_documents').where('batchId', '==', batch_id).stream()
    result = []
    for doc in docs:
        data = doc.to_dict()
        data['id'] = doc.id
        result.append(data)
    result.sort(key=lambda x: x.get('uploadedAt', ''), reverse=True)
    return result

@api_router.post("/batches/{batch_id}/documents")
async def upload_batch_document(batch_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    batch_doc = db.collection('batches').document(batch_id).get()
    if not batch_doc.exists:
        raise HTTPException(status_code=404, detail="Batch not found")

    data = await file.read()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large (max 10MB)")

    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
    file_id = str(uuid.uuid4())
    storage_path = f"{APP_NAME}/batches/{batch_id}/documents/{file_id}.{ext}"

    try:
        result = put_object(storage_path, data, file.content_type or "application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

    doc_data = {
        "batchId": batch_id,
        "fileId": file_id,
        "name": file.filename,
        "storagePath": result["path"],
        "contentType": file.content_type or "application/octet-stream",
        "size": len(data),
        "uploadedBy": user.get("displayName", "Unknown"),
        "uploadedByUid": user["uid"],
        "uploadedAt": datetime.now(timezone.utc).isoformat()
    }

    ref = db.collection('batch_documents').document()
    ref.set(doc_data)
    doc_data['id'] = ref.id
    return doc_data

@api_router.get("/batches/{batch_id}/documents/{doc_id}/download")
async def download_batch_document(batch_id: str, doc_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    doc = db.collection('batch_documents').document(doc_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Document not found")
    doc_data = doc.to_dict()
    try:
        content, content_type = get_object(doc_data["storagePath"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")
    return Response(
        content=content,
        media_type=doc_data.get("contentType", content_type),
        headers={"Content-Disposition": f'attachment; filename="{doc_data["name"]}"'}
    )

@api_router.delete("/batches/{batch_id}/documents/{doc_id}")
async def delete_batch_document(batch_id: str, doc_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    db.collection('batch_documents').document(doc_id).delete()
    return {"message": "Document deleted"}

# Batch Feedback Routes
@api_router.get("/batches/{batch_id}/feedback")
async def list_batch_feedback(batch_id: str, user: dict = Depends(get_current_user)):
    if not db:
        return []
    docs = db.collection('batch_feedback').where('batchId', '==', batch_id).stream()
    result = []
    for doc in docs:
        data = doc.to_dict()
        data['id'] = doc.id
        result.append(data)
    result.sort(key=lambda x: x.get('updatedAt', x.get('createdAt', '')), reverse=True)
    return result

@api_router.post("/batches/{batch_id}/feedback")
async def create_or_update_feedback(batch_id: str, feedback_data: dict, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    batch_doc = db.collection('batches').document(batch_id).get()
    if not batch_doc.exists:
        raise HTTPException(status_code=404, detail="Batch not found")

    existing = list(db.collection('batch_feedback').where('batchId', '==', batch_id).where('userId', '==', user['uid']).limit(1).stream())

    clean = {
        'batchId': batch_id,
        'userId': user['uid'],
        'leadName': user.get('displayName', 'Unknown'),
        'positive': str(feedback_data.get('positive', '')).strip(),
        'negative': str(feedback_data.get('negative', '')).strip(),
        'updatedAt': datetime.now(timezone.utc).isoformat()
    }

    if existing:
        doc_id = existing[0].id
        db.collection('batch_feedback').document(doc_id).update(clean)
        clean['id'] = doc_id
    else:
        clean['createdAt'] = datetime.now(timezone.utc).isoformat()
        ref = db.collection('batch_feedback').document()
        ref.set(clean)
        clean['id'] = ref.id

    return clean

# Lead Routes
@api_router.post("/leads")
async def create_lead(lead: LeadCreate, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    lead_data = lead.model_dump()
    email = lead_data.pop('email', None)
    password = lead_data.pop('password', None)
    role = lead_data.pop('role', 'Trek Lead')
    
    lead_data['createdAt'] = datetime.now(timezone.utc).isoformat()
    lead_data['assignedBatches'] = []
    
    # If email & password provided, create Firebase Auth user + Firestore user
    user_uid = None
    if email and password:
        try:
            user_record = firebase_auth.create_user(
                email=email,
                password=password,
                display_name=lead_data['name']
            )
            user_uid = user_record.uid
            # Create Firestore user doc (auto-approved since admin is creating)
            user_doc = {
                'email': email,
                'displayName': lead_data['name'],
                'role': role,
                'status': 'approved',
                'createdAt': datetime.now(timezone.utc).isoformat()
            }
            db.collection('users').document(user_uid).set(user_doc)
            lead_data['userId'] = user_uid
            lead_data['email'] = email
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to create user account: {str(e)}")
    
    doc_ref = db.collection('leads').document()
    doc_ref.set(lead_data)
    
    lead_data['id'] = doc_ref.id
    return lead_data

@api_router.get("/leads")
async def list_leads(approval_status: Optional[str] = Query(None), user: dict = Depends(get_current_user)):
    if not db:
        return []

    leads_ref = db.collection('leads')

    if approval_status:
        leads_docs = await fs_run(lambda: list(leads_ref.where('approvalStatus', '==', approval_status).stream()), default=[])
    else:
        # Default: only show approved leads (or legacy leads without approvalStatus field)
        leads_docs = await fs_run(lambda: list(leads_ref.stream()), default=[])

    lead_list = []
    for lead_doc in leads_docs:
        lead_data = lead_doc.to_dict()
        # Exclude pending/rejected unless explicitly requested
        if not approval_status and lead_data.get('approvalStatus') == 'pending':
            continue
        if not approval_status and lead_data.get('approvalStatus') == 'rejected':
            continue
        lead_data['id'] = lead_doc.id
        lead_list.append(lead_data)

    return lead_list

@api_router.patch("/leads/{lead_id}")
async def update_lead(lead_id: str, lead_data: dict, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    lead_ref = db.collection('leads').document(lead_id)
    lead_data['updatedAt'] = datetime.now(timezone.utc).isoformat()
    lead_ref.update(lead_data)
    
    return {'message': 'Lead updated successfully', 'id': lead_id}

@api_router.post("/leads/register")
async def self_register_lead(lead: LeadSelfRegister):
    """Public endpoint — trek lead candidates apply themselves; requires admin approval."""
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")

    # Create Firebase Auth account
    try:
        user_record = firebase_auth.create_user(
            email=lead.email,
            password=lead.password,
            display_name=lead.name
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to create account: {str(e)}")

    user_uid = user_record.uid

    # Firestore user doc — pending until admin approves
    db.collection('users').document(user_uid).set({
        'email': lead.email,
        'displayName': lead.name,
        'role': 'Trek Lead',
        'status': 'pending',
        'createdAt': datetime.now(timezone.utc).isoformat()
    })

    # Firestore lead doc — approvalStatus: pending
    lead_data = {
        'name': lead.name,
        'phone': lead.phone,
        'age': lead.age,
        'gender': lead.gender,
        'hiredDate': lead.hiredDate,
        'languages': lead.languages,
        'specialSkills': lead.specialSkills,
        'email': lead.email,
        'active': False,
        'approvalStatus': 'pending',
        'userId': user_uid,
        'assignedBatches': [],
        'createdAt': datetime.now(timezone.utc).isoformat()
    }
    doc_ref = db.collection('leads').document()
    doc_ref.set(lead_data)

    return {'message': 'Application submitted successfully. Awaiting admin approval.', 'id': doc_ref.id}

@api_router.post("/leads/{lead_id}/approve")
async def approve_lead(lead_id: str, user: dict = Depends(require_admin)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")

    lead_ref = db.collection('leads').document(lead_id)
    lead_doc = lead_ref.get()
    if not lead_doc.exists:
        raise HTTPException(status_code=404, detail="Lead not found")

    lead_ref.update({'approvalStatus': 'approved', 'active': True, 'updatedAt': datetime.now(timezone.utc).isoformat()})

    # Also approve the linked user account
    lead_data = lead_doc.to_dict()
    user_id = lead_data.get('userId')
    if user_id:
        db.collection('users').document(user_id).update({'status': 'approved'})

    return {'message': 'Lead approved successfully', 'id': lead_id}

@api_router.post("/leads/{lead_id}/reject")
async def reject_lead(lead_id: str, user: dict = Depends(require_admin)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")

    lead_ref = db.collection('leads').document(lead_id)
    lead_doc = lead_ref.get()
    if not lead_doc.exists:
        raise HTTPException(status_code=404, detail="Lead not found")

    lead_data = lead_doc.to_dict()
    user_id = lead_data.get('userId')

    # Delete Firebase Auth user and Firestore user doc
    if user_id:
        try:
            firebase_auth.delete_user(user_id)
        except Exception:
            pass
        db.collection('users').document(user_id).delete()

    lead_ref.delete()
    return {'message': 'Lead application rejected and removed', 'id': lead_id}

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, user: dict = Depends(require_admin)):
    """Permanently delete an approved lead profile and their associated user account."""
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")

    lead_ref = db.collection('leads').document(lead_id)
    lead_doc = lead_ref.get()
    if not lead_doc.exists:
        raise HTTPException(status_code=404, detail="Lead not found")

    lead_data = lead_doc.to_dict()
    user_id = lead_data.get('userId')

    if user_id:
        try:
            firebase_auth.delete_user(user_id)
        except Exception:
            pass
        db.collection('users').document(user_id).delete()

    lead_ref.delete()
    return {'message': 'Lead profile deleted permanently', 'id': lead_id}

# ── Payout Routes (reads finance Firestore) ────────────────────────────────

@api_router.get("/leads/{lead_id}/payouts")
async def get_lead_payouts(lead_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    lead_doc = await fs_run(lambda: db.collection('leads').document(lead_id).get())
    if not lead_doc or not lead_doc.exists:
        raise HTTPException(status_code=404, detail="Lead not found")
    lead_data = lead_doc.to_dict()
    lead_name = (lead_data.get('name') or '').strip()
    if not finance_db:
        return {'leadId': lead_id, 'leadName': lead_name, 'entries': [],
                'totalBatches': 0, 'totalEarned': 0, 'financeConnected': False}
    batches_raw = await fs_run(lambda: list(finance_db.collection('batches').stream()), default=[])
    entries = []
    for b_doc in batches_raw:
        b = b_doc.to_dict()
        b['id'] = b_doc.id
        lp_array = b.get('leadPayments', [])
        amount_from_array = sum(int(lp.get('amount') or 0) for lp in lp_array
                                if (lp.get('name') or '').strip().lower() == lead_name.lower())
        legacy_hit = ((b.get('leadId') and b['leadId'] == lead_id) or
                      (b.get('leadName') or '').strip().lower() == lead_name.lower())
        amount_from_legacy = int(b.get('leadPayment') or 0) if legacy_hit else 0
        total_amount = amount_from_array + amount_from_legacy
        if total_amount <= 0:
            continue
        cleared_by = b.get('paymentClearedBy') or {}
        if lead_id in cleared_by and isinstance(cleared_by[lead_id], bool):
            cleared = cleared_by[lead_id]
        elif lead_name in cleared_by and isinstance(cleared_by[lead_name], bool):
            cleared = cleared_by[lead_name]
        else:
            cleared = bool(b.get('paymentCleared', False))
        lead_key = lead_id if lead_id in cleared_by else lead_name
        entries.append({'financeBatchId': b['id'], 'batchCode': b.get('batchCode') or '—',
                        'date': b.get('date') or b.get('startDate') or '—',
                        'trekName': b.get('trekName') or '—', 'amount': total_amount,
                        'cleared': cleared, 'leadKey': lead_key})
    entries.sort(key=lambda e: e['date'], reverse=True)
    return {'leadId': lead_id, 'leadName': lead_name, 'entries': entries,
            'totalBatches': len(entries), 'totalEarned': sum(e['amount'] for e in entries),
            'financeConnected': True}

@api_router.patch("/leads/{lead_id}/payouts/{finance_batch_id}/toggle")
async def toggle_lead_payout(lead_id: str, finance_batch_id: str, body: dict,
                              user: dict = Depends(require_admin)):
    if not finance_db:
        raise HTTPException(status_code=503, detail="Finance database not connected")
    lead_key = body.get('leadKey')
    new_status = body.get('cleared')
    if lead_key is None or new_status is None:
        raise HTTPException(status_code=400, detail="leadKey and cleared are required")
    batch_ref = finance_db.collection('batches').document(finance_batch_id)
    batch_doc = await fs_run(lambda: batch_ref.get())
    if not batch_doc or not batch_doc.exists:
        raise HTTPException(status_code=404, detail="Finance batch not found")
    batch_ref.update({f'paymentClearedBy.{lead_key}': bool(new_status)})
    return {'message': 'Payment status updated', 'leadKey': lead_key, 'cleared': new_status}

# Checklist Routes
@api_router.post("/checklists")
async def create_checklist(checklist: ChecklistCreate, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    checklist_data = checklist.model_dump()
    checklist_data['createdAt'] = datetime.now(timezone.utc).isoformat()
    
    doc_ref = db.collection('checklists').document()
    doc_ref.set(checklist_data)
    
    checklist_data['id'] = doc_ref.id
    return checklist_data

@api_router.get("/checklists/{batch_id}")
async def get_checklists(batch_id: str, user: dict = Depends(get_current_user)):
    if not db:
        return []
    
    checklists = db.collection('checklists').where('batchId', '==', batch_id).stream()
    
    checklist_list = []
    for checklist_doc in checklists:
        checklist_data = checklist_doc.to_dict()
        checklist_data['id'] = checklist_doc.id
        checklist_list.append(checklist_data)
    
    return checklist_list

@api_router.patch("/checklists/{checklist_id}")
async def update_checklist(checklist_id: str, checklist_data: dict, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    checklist_ref = db.collection('checklists').document(checklist_id)
    checklist_data['updatedAt'] = datetime.now(timezone.utc).isoformat()
    checklist_ref.update(checklist_data)
    
    return {'message': 'Checklist updated successfully', 'id': checklist_id}

# Dashboard Stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    if not db:
        return {
            'totalUpcomingBatches': 0,
            'totalActiveLeads': 0,
            'totalActiveTreks': 0,
            'completedBatchesThisMonth': 0,
            'pendingUsers': 0
        }
    
    try:
        user_role = user.get('role', '')
        is_admin = user_role in ('Super Admin', 'Operations Manager')
        today = datetime.now(timezone.utc).date().isoformat()
        
        if is_admin:
            upcoming_batches = await fs_run(lambda: list(db.collection('batches').where('startDate', '>=', today).stream()), default=[])
            active_leads = await fs_run(lambda: list(db.collection('leads').where('active', '==', True).stream()), default=[])
            active_treks = await fs_run(lambda: list(db.collection('treks').where('archived', '==', False).stream()), default=[])
            from datetime import datetime as dt
            first_day_of_month = dt.now(timezone.utc).replace(day=1).date().isoformat()
            all_completed = await fs_run(lambda: list(db.collection('batches').where('status', '==', 'Completed').stream()), default=[])
            completed_batches = [b for b in all_completed if b.to_dict().get('endDate', '') >= first_day_of_month]
            pending_list = await fs_run(lambda: list(db.collection('users').where('status', '==', 'pending').stream()), default=[]) if user_role == 'Super Admin' else []
            pending_users = len(pending_list)
        else:
            # For leads/coordinators - only their assigned batches
            all_batches = await fs_run(lambda: list(db.collection('batches').stream()), default=[])
            my_batches = []
            for b in all_batches:
                bd = b.to_dict()
                assigned = bd.get('assignedLeads', [])
                if any(a.get('userId') == user['uid'] for a in assigned):
                    my_batches.append(bd)
            upcoming_batches = [b for b in my_batches if b.get('startDate', '') >= today]
            from datetime import datetime as dt
            first_day_of_month = dt.now(timezone.utc).replace(day=1).date().isoformat()
            completed_batches = [b for b in my_batches if b.get('status') == 'Completed' and b.get('endDate', '') >= first_day_of_month]
            active_leads = []
            active_treks = []
            pending_users = 0
        
        return {
            'totalUpcomingBatches': len(upcoming_batches),
            'totalActiveLeads': len(active_leads),
            'totalActiveTreks': len(active_treks),
            'completedBatchesThisMonth': len(completed_batches),
            'pendingUsers': pending_users
        }
    except Exception as e:
        logger.error(f"Error fetching dashboard stats: {str(e)}")
        return {
            'totalUpcomingBatches': 0,
            'totalActiveLeads': 0,
            'totalActiveTreks': 0,
            'completedBatchesThisMonth': 0,
            'pendingUsers': 0
        }

# Ticket Routes
@api_router.post("/tickets")
async def create_ticket(ticket: TicketCreate, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    ticket_data = ticket.model_dump()
    ticket_data['createdAt'] = datetime.now(timezone.utc).isoformat()
    ticket_data['createdBy'] = user['uid']
    ticket_data['reporter'] = user['displayName']
    ticket_data['comments'] = []
    ticket_data['activityLog'] = [{
        'action': 'created',
        'user': user['displayName'],
        'timestamp': datetime.now(timezone.utc).isoformat()
    }]
    
    doc_ref = db.collection('tickets').document()
    doc_ref.set(ticket_data)
    
    ticket_data['id'] = doc_ref.id
    return ticket_data

@api_router.get("/tickets")
async def list_tickets(user: dict = Depends(get_current_user), 
                       status: Optional[str] = None, 
                       category: Optional[str] = None,
                       assignee: Optional[str] = None,
                       priority: Optional[str] = None,
                       search: Optional[str] = None,
                       start_date: Optional[str] = None,
                       end_date: Optional[str] = None,
                       page: int = 1,
                       limit: int = 100):
    if not db:
        return {'tickets': [], 'total': 0, 'page': page, 'pages': 1}
    
    tickets_ref = db.collection('tickets')
    
    # Apply basic filters (Firestore limitations - can't do complex queries)
    # We'll filter in Python for more complex queries
    all_tickets = list(tickets_ref.stream())
    
    # Role-based filtering: non-admin users only see their assigned tasks
    user_role = user.get('role', '')
    is_admin = user_role in ('Super Admin', 'Operations Manager')
    
    # Filter in Python for flexibility
    filtered_tickets = []
    for ticket_doc in all_tickets:
        ticket_data = ticket_doc.to_dict()
        ticket_data['id'] = ticket_doc.id
        
        # Non-admin users only see tasks assigned to them or created by them
        if not is_admin:
            assignees = ticket_data.get('assignees', [])
            if user['uid'] not in assignees and ticket_data.get('createdBy') != user['uid']:
                continue
        
        # Apply filters
        if status and ticket_data.get('status') != status:
            continue
        if category and ticket_data.get('category') != category:
            continue
        if priority and ticket_data.get('priority') != priority:
            continue
        if assignee:
            assignees = ticket_data.get('assignees', [])
            if assignee not in assignees:
                continue
        if search:
            search_lower = search.lower()
            title = ticket_data.get('title', '').lower()
            description = ticket_data.get('description', '').lower()
            if search_lower not in title and search_lower not in description:
                continue
        if start_date:
            ticket_date = ticket_data.get('createdAt', '')
            if ticket_date < start_date:
                continue
        if end_date:
            ticket_date = ticket_data.get('createdAt', '')
            if ticket_date > end_date:
                continue
        
        filtered_tickets.append(ticket_data)
    
    # Sort by created date (newest first)
    filtered_tickets.sort(key=lambda x: x.get('createdAt', ''), reverse=True)
    
    # Pagination
    total = len(filtered_tickets)
    pages = (total + limit - 1) // limit
    start_idx = (page - 1) * limit
    end_idx = start_idx + limit
    paginated_tickets = filtered_tickets[start_idx:end_idx]
    
    return {
        'tickets': paginated_tickets,
        'total': total,
        'page': page,
        'pages': pages,
        'limit': limit
    }

@api_router.get("/tickets/{ticket_id}")
async def get_ticket(ticket_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    ticket_doc = db.collection('tickets').document(ticket_id).get()
    
    if not ticket_doc.exists:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    ticket_data = ticket_doc.to_dict()
    ticket_data['id'] = ticket_doc.id
    return ticket_data

@api_router.patch("/tickets/{ticket_id}")
async def update_ticket(ticket_id: str, ticket_update: TicketUpdate, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    ticket_ref = db.collection('tickets').document(ticket_id)
    ticket_doc = ticket_ref.get()
    
    if not ticket_doc.exists:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    update_data = {k: v for k, v in ticket_update.model_dump().items() if v is not None}
    update_data['updatedAt'] = datetime.now(timezone.utc).isoformat()
    
    # Add to activity log
    current_ticket = ticket_doc.to_dict()
    activity_log = current_ticket.get('activityLog', [])
    
    def strip_html(text):
        return re.sub(r'<[^>]+>', '', str(text)).strip()

    for key, value in update_data.items():
        if key in current_ticket and current_ticket[key] != value:
            activity_log.append({
                'action': f'updated {key}',
                'from': strip_html(current_ticket[key]),
                'to': strip_html(value),
                'user': user['displayName'],
                'timestamp': datetime.now(timezone.utc).isoformat()
            })
    
    update_data['activityLog'] = activity_log
    
    ticket_ref.update(update_data)
    
    return {'message': 'Ticket updated successfully', 'id': ticket_id}

@api_router.post("/tickets/{ticket_id}/comments")
async def add_comment(ticket_id: str, comment_data: dict, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    ticket_ref = db.collection('tickets').document(ticket_id)
    ticket_doc = ticket_ref.get()
    
    if not ticket_doc.exists:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    ticket_data = ticket_doc.to_dict()
    comments = ticket_data.get('comments', [])
    
    new_comment = {
        'text': comment_data.get('text', ''),
        'user': user['displayName'],
        'userId': user['uid'],
        'timestamp': datetime.now(timezone.utc).isoformat()
    }
    
    comments.append(new_comment)
    ticket_ref.update({'comments': comments})
    
    return {'message': 'Comment added successfully', 'comment': new_comment}

@api_router.delete("/tickets/{ticket_id}")
async def delete_ticket(ticket_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    db.collection('tickets').document(ticket_id).delete()
    return {'message': 'Ticket deleted successfully'}

# File Attachment Routes
@api_router.post("/tickets/{ticket_id}/attachments")
async def upload_attachment(ticket_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")

    ticket_ref = db.collection('tickets').document(ticket_id)
    ticket_doc = ticket_ref.get()
    if not ticket_doc.exists:
        raise HTTPException(status_code=404, detail="Ticket not found")

    data = await file.read()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large (max 10MB)")

    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
    file_id = str(uuid.uuid4())
    storage_path = f"{APP_NAME}/tickets/{ticket_id}/{file_id}.{ext}"

    try:
        result = put_object(storage_path, data, file.content_type or "application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

    attachment = {
        "id": file_id,
        "name": file.filename,
        "storagePath": result["path"],
        "contentType": file.content_type or "application/octet-stream",
        "size": len(data),
        "uploadedBy": user.get("displayName", "Unknown"),
        "uploadedAt": datetime.now(timezone.utc).isoformat()
    }

    ticket_data = ticket_doc.to_dict()
    attachments = ticket_data.get("attachments", [])
    attachments.append(attachment)
    ticket_ref.update({"attachments": attachments})

    return attachment

@api_router.get("/attachments/{file_id}")
async def download_attachment(file_id: str, ticket_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")

    ticket_doc = db.collection('tickets').document(ticket_id).get()
    if not ticket_doc.exists:
        raise HTTPException(status_code=404, detail="Ticket not found")

    ticket_data = ticket_doc.to_dict()
    attachments = ticket_data.get("attachments", [])
    attachment = next((a for a in attachments if a["id"] == file_id), None)
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")

    try:
        content, content_type = get_object(attachment["storagePath"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")

    return Response(
        content=content,
        media_type=attachment.get("contentType", content_type),
        headers={"Content-Disposition": f'attachment; filename="{attachment["name"]}"'}
    )

@api_router.delete("/tickets/{ticket_id}/attachments/{file_id}")
async def delete_attachment(ticket_id: str, file_id: str, user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")

    ticket_ref = db.collection('tickets').document(ticket_id)
    ticket_doc = ticket_ref.get()
    if not ticket_doc.exists:
        raise HTTPException(status_code=404, detail="Ticket not found")

    ticket_data = ticket_doc.to_dict()
    attachments = ticket_data.get("attachments", [])
    attachments = [a for a in attachments if a["id"] != file_id]
    ticket_ref.update({"attachments": attachments})

    return {"message": "Attachment removed"}

# Workload Analytics
@api_router.get("/tickets/analytics/workload")
async def get_workload_analytics(user: dict = Depends(get_current_user), search: Optional[str] = None):
    if not db:
        return []
    
    # Get all active tickets (not Done)
    active_tickets = db.collection('tickets').where('status', '!=', 'Done').stream()
    
    # Group by assignee (handle multiple assignees)
    workload = {}
    for ticket_doc in active_tickets:
        ticket_data = ticket_doc.to_dict()
        assignees = ticket_data.get('assignees', ['Unassigned'])
        
        if not assignees:
            assignees = ['Unassigned']
        
        for assignee in assignees:
            if assignee not in workload:
                workload[assignee] = {
                    'assignee': assignee,
                    'totalTickets': 0,
                    'byPriority': {'Low': 0, 'Medium': 0, 'High': 0, 'Urgent': 0},
                    'byStatus': {},
                    'byCategory': {},
                    'totalEstimatedHours': 0,
                    'tickets': []
                }
            
            workload[assignee]['totalTickets'] += 1
            priority = ticket_data.get('priority', 'Medium')
            workload[assignee]['byPriority'][priority] = workload[assignee]['byPriority'].get(priority, 0) + 1
            
            status = ticket_data.get('status', 'Backlog')
            workload[assignee]['byStatus'][status] = workload[assignee]['byStatus'].get(status, 0) + 1
            
            category = ticket_data.get('category', 'Operations')
            workload[assignee]['byCategory'][category] = workload[assignee]['byCategory'].get(category, 0) + 1
            
            estimated_hours = ticket_data.get('estimatedHours', 0)
            if estimated_hours:
                workload[assignee]['totalEstimatedHours'] += estimated_hours
            
            # Add ticket summary
            workload[assignee]['tickets'].append({
                'id': ticket_doc.id,
                'title': ticket_data.get('title', ''),
                'priority': priority,
                'status': status,
                'dueDate': ticket_data.get('dueDate', '')
            })
    
    result = list(workload.values())
    
    # Apply search filter
    if search:
        search_lower = search.lower()
        result = [w for w in result if search_lower in w['assignee'].lower()]
    
    return result

# Config Routes
@api_router.get("/config")
async def get_config(user: dict = Depends(get_current_user)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    config_doc = db.collection('config').document('dropdown_options').get()
    
    if config_doc.exists:
        config_data = config_doc.to_dict()
        return config_data
    else:
        # Return defaults if no config exists
        default_config = {
            'categories': ['Karnataka', 'Kerala', 'Himalayas', 'Sunrise', 'Backpacking', 'Kids Batch'],
            'difficultyLevels': ['Easy', 'Moderate', 'Difficult', 'Very Difficult'],
            'trekTypes': ['1-day', '2-day', 'Himalayan'],
            'batchStatuses': ['Open', 'Filling Fast', 'Full', 'Closed', 'Completed', 'Cancelled'],
            'taskCategories': ['Operations', 'Sales', 'Content', 'Development', 'Trek Planning']
        }
        return default_config

@api_router.post("/config")
async def save_config(config_data: dict, admin: dict = Depends(require_admin)):
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    config_ref = db.collection('config').document('dropdown_options')
    config_ref.set(config_data)
    
    return {'message': 'Configuration saved successfully'}

# -------------------------------------------------------
# Badge / Voucher System
# -------------------------------------------------------

BADGE_TIERS = [
    {"id": "kumara_parvatha", "name": "Kumara Parvatha", "elevation": "1,712m", "minBatches": 5,  "emoji": "🏔️"},
    {"id": "kedarkantha",     "name": "Kedarkantha",     "elevation": "3,810m", "minBatches": 10, "emoji": "⛰️"},
    {"id": "roopkund",        "name": "Roopkund",        "elevation": "5,029m", "minBatches": 20, "emoji": "🗻"},
    {"id": "trishul",         "name": "Trishul",         "elevation": "7,120m", "minBatches": 30, "emoji": "🌟"},
    {"id": "nanda_devi",      "name": "Nanda Devi",      "elevation": "7,816m", "minBatches": 40, "emoji": "💎"},
    {"id": "everester",       "name": "Everester",       "elevation": "8,849m", "minBatches": 50, "emoji": "🏆"},
]

@api_router.get("/badge-config")
async def get_badge_config(user: dict = Depends(get_current_user)):
    """Return all badge tiers with admin-configured goodie info."""
    tiers = []
    for tier in BADGE_TIERS:
        config = {}
        if db:
            doc = await fs_run(
                lambda tid=tier["id"]: db.collection("badge_config").document(tid).get()
            )
            if doc and doc.exists:
                config = doc.to_dict()
        tiers.append({**tier, **config})
    return tiers

@api_router.put("/badge-config/{tier_id}")
async def update_badge_config(tier_id: str, data: dict, admin: dict = Depends(require_admin)):
    """Admin: set goodie description (and optionally pic URL) for a tier."""
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    valid_ids = {t["id"] for t in BADGE_TIERS}
    if tier_id not in valid_ids:
        raise HTTPException(status_code=404, detail="Unknown tier")
    allowed = {"goodieDescription", "goodiePicUrl"}
    payload = {k: v for k, v in data.items() if k in allowed}
    db.collection("badge_config").document(tier_id).set(payload, merge=True)
    return {"message": "Badge config updated"}

@api_router.post("/badge-config/{tier_id}/upload-image")
async def upload_badge_image(tier_id: str, file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    """Admin: upload a goodie photo for a tier."""
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    valid_ids = {t["id"] for t in BADGE_TIERS}
    if tier_id not in valid_ids:
        raise HTTPException(status_code=404, detail="Unknown tier")
    data = await file.read()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large (max 10MB)")
    ext = Path(file.filename).suffix or ".jpg"
    path = f"badge-images/{tier_id}{ext}"
    put_object(path, data, file.content_type or "image/jpeg")
    pic_url = f"/api/badge-config/{tier_id}/image"
    db.collection("badge_config").document(tier_id).set({"goodiePicUrl": pic_url}, merge=True)
    return {"goodiePicUrl": pic_url}

@api_router.get("/badge-config/{tier_id}/image")
async def get_badge_image(tier_id: str):
    """Serve the uploaded goodie image for a tier."""
    for ext in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
        path = f"badge-images/{tier_id}{ext}"
        try:
            content, content_type = get_object(path)
            return Response(content=content, media_type=content_type)
        except Exception:
            continue
    raise HTTPException(status_code=404, detail="No image found for this tier")

@api_router.post("/badges/claim/{tier_id}")
async def claim_badge(tier_id: str, user: dict = Depends(get_current_user)):
    """Lead claims an unlocked badge. Generates a unique voucher code."""
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
    uid = user["uid"]
    tier = next((t for t in BADGE_TIERS if t["id"] == tier_id), None)
    if not tier:
        raise HTTPException(status_code=404, detail="Unknown badge tier")

    # Check if already claimed
    voucher_id = f"{uid}_{tier_id}"
    existing = await fs_run(
        lambda: db.collection("vouchers").document(voucher_id).get()
    )
    if existing and existing.exists:
        return existing.to_dict()

    # Verify the lead has enough completed batches
    # (array-contains on nested objects doesn't work in Firestore — stream all and filter)
    all_batches = await fs_run(
        lambda: list(db.collection("batches").stream()),
        timeout=8.0,
        default=[]
    )
    completed = 0
    for b in all_batches:
        bd = b.to_dict()
        if bd.get("status") == "Completed":
            if any(l.get("userId") == uid for l in bd.get("assignedLeads", [])):
                completed += 1

    if completed < tier["minBatches"]:
        raise HTTPException(
            status_code=403,
            detail=f"Not enough completed batches. Need {tier['minBatches']}, have {completed}."
        )

    # Fetch goodie config
    config_doc = await fs_run(
        lambda: db.collection("badge_config").document(tier_id).get()
    )
    config = config_doc.to_dict() if (config_doc and config_doc.exists) else {}

    # Generate voucher
    code = f"BT-{tier_id.upper()[:3]}-{str(uuid.uuid4())[:8].upper()}"
    voucher = {
        "userId": uid,
        "displayName": user.get("displayName", ""),
        "tierId": tier_id,
        "tierName": tier["name"],
        "elevation": tier["elevation"],
        "emoji": tier["emoji"],
        "voucherCode": code,
        "goodieDescription": config.get("goodieDescription", ""),
        "goodiePicUrl": config.get("goodiePicUrl", ""),
        "claimedAt": datetime.now(timezone.utc).isoformat(),
    }
    db.collection("vouchers").document(voucher_id).set(voucher)
    return voucher

@api_router.get("/badges/vouchers")
async def get_my_vouchers(user: dict = Depends(get_current_user)):
    """Lead fetches all their claimed vouchers."""
    if not db:
        return []
    uid = user["uid"]
    docs = await fs_run(
        lambda: db.collection("vouchers").where("userId", "==", uid).get(),
        timeout=6.0,
        default=[]
    )
    if not docs:
        return []
    return [d.to_dict() for d in docs]

@api_router.get("/badges/vouchers/all")
async def get_all_vouchers(admin: dict = Depends(require_admin)):
    """Admin: get all claimed vouchers across all leads."""
    if not db:
        return []
    docs = await fs_run(
        lambda: db.collection("vouchers").order_by("claimedAt", direction=firestore.Query.DESCENDING).get(),
        timeout=8.0,
        default=[]
    )
    if not docs:
        return []
    return [d.to_dict() for d in docs]

# Include the router in the main app
app.include_router(api_router)

@app.on_event("startup")
async def startup():
    UPLOAD_DIR.mkdir(exist_ok=True)
    logging.info(f"File storage initialized at {UPLOAD_DIR}")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8080))
    uvicorn.run("server:app", host="0.0.0.0", port=port)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)