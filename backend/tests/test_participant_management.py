"""
Test suite for Phase 2 Participant Management feature
- GET/POST /api/batches/{batch_id}/participants 
- PATCH/DELETE /api/batches/{batch_id}/participants/{pid}
- GET /api/batches/{batch_id}/participants/template (xlsx download)
- POST /api/batches/{batch_id}/participants/import (xlsx import)
"""
import pytest
import requests
import os
import time
import io
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
ADMIN_EMAIL = "admin@bengalurutrekkers.com"
ADMIN_PASSWORD = "admin123456"


class TestParticipantManagement:
    """Test participant CRUD and Excel import/export functionality"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Get auth token before each test"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Get Firebase auth token
        firebase_api_key = "AIzaSyCFl3n1PmSTN6eO4NdNDevIvFAmO78fV_M"
        auth_url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={firebase_api_key}"
        
        auth_response = self.session.post(auth_url, json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD,
            "returnSecureToken": True
        })
        
        if auth_response.status_code == 200:
            self.token = auth_response.json().get("idToken")
            self.user_id = auth_response.json().get("localId")
            self.session.headers.update({"Authorization": f"Bearer {self.token}"})
        else:
            pytest.skip(f"Authentication failed: {auth_response.text}")
    
    def get_or_create_test_batch(self):
        """Get existing test batch or create one for testing"""
        # Get existing batches
        response = self.session.get(f"{BASE_URL}/api/batches")
        assert response.status_code == 200, f"Failed to get batches: {response.text}"
        batches = response.json()
        
        # Return first available batch
        if len(batches) > 0:
            return batches[0]
        
        # Create a new batch if none exist
        treks_response = self.session.get(f"{BASE_URL}/api/treks")
        assert treks_response.status_code == 200
        treks = treks_response.json()
        
        if len(treks) == 0:
            pytest.skip("No treks available for batch creation")
        
        batch_code = f"TEST_PART_{int(time.time())}"
        batch_data = {
            "batchCode": batch_code,
            "trekId": treks[0]["id"],
            "startDate": "2026-04-01",
            "endDate": "2026-04-03",
            "maxCapacity": 30,
            "currentRegistrations": 0,
            "status": "Open",
            "assignedLeads": []
        }
        
        create_response = self.session.post(f"{BASE_URL}/api/batches", json=batch_data)
        assert create_response.status_code == 200, f"Failed to create batch: {create_response.text}"
        return create_response.json()
    
    # === HEALTH CHECK ===
    def test_health_endpoint(self):
        """Test API health check"""
        response = self.session.get(f"{BASE_URL}/api/health")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "healthy"
        assert data["firebase_initialized"] == True
        assert data["firestore_connected"] == True
        print("✓ Health endpoint working, Firebase connected")
    
    # === BATCH DETAIL ENDPOINT ===
    def test_get_single_batch(self):
        """Test GET /api/batches/{batch_id} endpoint"""
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        
        response = self.session.get(f"{BASE_URL}/api/batches/{batch_id}")
        assert response.status_code == 200
        data = response.json()
        
        assert "id" in data
        assert "batchCode" in data
        assert "status" in data
        assert "startDate" in data
        assert "endDate" in data
        assert "maxCapacity" in data
        
        print(f"✓ GET /api/batches/{batch_id} returned batch: {data['batchCode']}")
    
    # === PARTICIPANT LIST ===
    def test_list_participants_empty(self):
        """Test GET /api/batches/{batch_id}/participants endpoint"""
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        
        response = self.session.get(f"{BASE_URL}/api/batches/{batch_id}/participants")
        assert response.status_code == 200
        participants = response.json()
        assert isinstance(participants, list)
        print(f"✓ GET /api/batches/{batch_id}/participants returned {len(participants)} participants")
    
    # === TEMPLATE DOWNLOAD ===
    def test_download_participant_template(self):
        """Test GET /api/batches/{batch_id}/participants/template downloads xlsx"""
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        
        # Remove Content-Type for blob request
        headers = {"Authorization": f"Bearer {self.token}"}
        response = requests.get(
            f"{BASE_URL}/api/batches/{batch_id}/participants/template",
            headers=headers
        )
        
        assert response.status_code == 200, f"Template download failed: {response.text}"
        
        # Check content type
        content_type = response.headers.get("Content-Type", "")
        assert "spreadsheet" in content_type or "application/vnd.openxmlformats" in content_type, \
            f"Expected xlsx content type, got: {content_type}"
        
        # Check Content-Disposition header for filename
        content_disp = response.headers.get("Content-Disposition", "")
        assert "attachment" in content_disp
        assert ".xlsx" in content_disp
        
        # Verify file is a valid xlsx by loading it
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check header row
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            'SL No', 'Full Name', 'Contact No', 'Age', 'Gender', 'Pickup Point',
            'Total Price incl GST', 'Amount Paid', 'Balance Amount', 'Receipt Mode',
            'Receipt Date', 'If Cancelled Reason', 'Booked By', 'Remarks'
        ]
        assert headers == expected_headers, f"Headers mismatch: {headers}"
        
        print(f"✓ Template download works, contains {len(expected_headers)} headers")
    
    # === ADD PARTICIPANT ===
    def test_add_participant(self):
        """Test POST /api/batches/{batch_id}/participants"""
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        
        participant_data = {
            "slNo": "1",
            "fullName": "TEST_John Doe",
            "contactNo": "9876543210",
            "age": "28",
            "gender": "Male",
            "pickupPoint": "Majestic",
            "totalPrice": 4500.00,
            "amountPaid": 2000.00,
            "balanceAmount": 2500.00,
            "receiptMode": "UPI",
            "receiptDate": "2026-01-15",
            "bookedBy": "Admin",
            "remarks": "Test participant"
        }
        
        response = self.session.post(
            f"{BASE_URL}/api/batches/{batch_id}/participants",
            json=participant_data
        )
        assert response.status_code == 200, f"Add participant failed: {response.text}"
        
        created = response.json()
        assert created["fullName"] == "TEST_John Doe"
        assert created["batchId"] == batch_id
        assert "id" in created
        assert created["boarded"] == False
        assert created["status"] == "Confirmed"
        
        # Store for later tests
        self.__class__.test_participant_id = created["id"]
        self.__class__.test_batch_id = batch_id
        
        print(f"✓ Created participant: {created['fullName']} with id {created['id']}")
        return created
    
    # === UPDATE PARTICIPANT ===
    def test_update_participant(self):
        """Test PATCH /api/batches/{batch_id}/participants/{pid}"""
        # Create participant first
        created = self.test_add_participant()
        batch_id = created["batchId"]
        participant_id = created["id"]
        
        # Update participant
        update_data = {
            "boarded": True,
            "amountCollected": 500.00,
            "remarks": "Updated remarks"
        }
        
        response = self.session.patch(
            f"{BASE_URL}/api/batches/{batch_id}/participants/{participant_id}",
            json=update_data
        )
        assert response.status_code == 200, f"Update failed: {response.text}"
        
        # Verify update by fetching participant list
        list_response = self.session.get(f"{BASE_URL}/api/batches/{batch_id}/participants")
        assert list_response.status_code == 200
        participants = list_response.json()
        
        updated = next((p for p in participants if p["id"] == participant_id), None)
        assert updated is not None, "Updated participant not found in list"
        assert updated["boarded"] == True
        assert updated["amountCollected"] == 500.00
        
        print(f"✓ Updated participant {participant_id}: boarded=True, amountCollected=500")
    
    # === DELETE PARTICIPANT ===
    def test_delete_participant(self):
        """Test DELETE /api/batches/{batch_id}/participants/{pid}"""
        # Create participant first
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        
        participant_data = {
            "fullName": "TEST_ToBeDeleted",
            "contactNo": "1234567890",
            "totalPrice": 1000.00,
            "amountPaid": 500.00,
            "balanceAmount": 500.00
        }
        
        create_response = self.session.post(
            f"{BASE_URL}/api/batches/{batch_id}/participants",
            json=participant_data
        )
        assert create_response.status_code == 200
        participant_id = create_response.json()["id"]
        
        # Delete participant
        delete_response = self.session.delete(
            f"{BASE_URL}/api/batches/{batch_id}/participants/{participant_id}"
        )
        assert delete_response.status_code == 200, f"Delete failed: {delete_response.text}"
        
        # Verify deletion
        list_response = self.session.get(f"{BASE_URL}/api/batches/{batch_id}/participants")
        participants = list_response.json()
        
        deleted = next((p for p in participants if p["id"] == participant_id), None)
        assert deleted is None, "Participant still exists after deletion"
        
        print(f"✓ Deleted participant {participant_id}")
    
    # === IMPORT EXCEL ===
    def test_import_participants_excel(self):
        """Test POST /api/batches/{batch_id}/participants/import"""
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Participants"
        
        # Headers
        headers = [
            'SL No', 'Full Name', 'Contact No', 'Age', 'Gender', 'Pickup Point',
            'Total Price incl GST', 'Amount Paid', 'Balance Amount', 'Receipt Mode',
            'Receipt Date', 'If Cancelled Reason', 'Booked By', 'Remarks'
        ]
        ws.append(headers)
        
        # Add test rows
        test_rows = [
            ['1', 'TEST_Import_User1', '9001234567', '25', 'Male', 'Silk Board', 5000, 2500, 2500, 'Cash', '2026-01-10', '', 'Online', 'Test import 1'],
            ['2', 'TEST_Import_User2', '9001234568', '30', 'Female', 'Koramangala', 5000, 5000, 0, 'UPI', '2026-01-11', '', 'Online', 'Test import 2'],
            ['3', 'TEST_Import_User3', '9001234569', '22', 'Male', 'HSR Layout', 4500, 2000, 2500, 'Card', '2026-01-12', '', 'Agent', 'Test import 3'],
        ]
        
        for row in test_rows:
            ws.append(row)
        
        # Save to bytes
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Upload file
        files = {
            'file': ('test_participants.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        headers = {"Authorization": f"Bearer {self.token}"}
        
        response = requests.post(
            f"{BASE_URL}/api/batches/{batch_id}/participants/import",
            headers=headers,
            files=files
        )
        
        assert response.status_code == 200, f"Import failed: {response.text}"
        
        result = response.json()
        assert result["count"] == 3, f"Expected 3 imported, got {result['count']}"
        assert "participants" in result
        
        # Verify imported participants
        imported_names = [p["fullName"] for p in result["participants"]]
        assert "TEST_Import_User1" in imported_names
        assert "TEST_Import_User2" in imported_names
        assert "TEST_Import_User3" in imported_names
        
        print(f"✓ Imported {result['count']} participants successfully")
        
        # Store IDs for cleanup
        self.__class__.imported_participant_ids = [p["id"] for p in result["participants"]]
        self.__class__.import_batch_id = batch_id
    
    # === BOARDING STATUS TOGGLE ===
    def test_toggle_boarding_status(self):
        """Test updating boarding status via PATCH"""
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        
        # Create a participant
        participant_data = {
            "fullName": "TEST_BoardingTest",
            "contactNo": "8888888888",
            "totalPrice": 3000.00,
            "amountPaid": 3000.00,
            "balanceAmount": 0
        }
        
        create_response = self.session.post(
            f"{BASE_URL}/api/batches/{batch_id}/participants",
            json=participant_data
        )
        assert create_response.status_code == 200
        participant = create_response.json()
        participant_id = participant["id"]
        
        # Verify initial boarded=False
        assert participant["boarded"] == False
        
        # Toggle to boarded=True
        toggle_response = self.session.patch(
            f"{BASE_URL}/api/batches/{batch_id}/participants/{participant_id}",
            json={"boarded": True}
        )
        assert toggle_response.status_code == 200
        
        # Verify
        list_response = self.session.get(f"{BASE_URL}/api/batches/{batch_id}/participants")
        participants = list_response.json()
        updated = next((p for p in participants if p["id"] == participant_id), None)
        assert updated["boarded"] == True
        
        # Toggle back to False
        toggle_response2 = self.session.patch(
            f"{BASE_URL}/api/batches/{batch_id}/participants/{participant_id}",
            json={"boarded": False}
        )
        assert toggle_response2.status_code == 200
        
        # Verify again
        list_response2 = self.session.get(f"{BASE_URL}/api/batches/{batch_id}/participants")
        participants2 = list_response2.json()
        updated2 = next((p for p in participants2 if p["id"] == participant_id), None)
        assert updated2["boarded"] == False
        
        print(f"✓ Boarding status toggle works correctly")
        
        # Cleanup
        self.session.delete(f"{BASE_URL}/api/batches/{batch_id}/participants/{participant_id}")
    
    # === AMOUNT COLLECTED UPDATE ===
    def test_update_amount_collected(self):
        """Test updating amountCollected via PATCH"""
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        
        # Create a participant with balance
        participant_data = {
            "fullName": "TEST_AmountTest",
            "contactNo": "7777777777",
            "totalPrice": 5000.00,
            "amountPaid": 2000.00,
            "balanceAmount": 3000.00
        }
        
        create_response = self.session.post(
            f"{BASE_URL}/api/batches/{batch_id}/participants",
            json=participant_data
        )
        assert create_response.status_code == 200
        participant_id = create_response.json()["id"]
        
        # Update amount collected
        update_response = self.session.patch(
            f"{BASE_URL}/api/batches/{batch_id}/participants/{participant_id}",
            json={"amountCollected": 1500.00}
        )
        assert update_response.status_code == 200
        
        # Verify
        list_response = self.session.get(f"{BASE_URL}/api/batches/{batch_id}/participants")
        participants = list_response.json()
        updated = next((p for p in participants if p["id"] == participant_id), None)
        assert updated["amountCollected"] == 1500.00
        
        print(f"✓ Amount collected update works correctly")
        
        # Cleanup
        self.session.delete(f"{BASE_URL}/api/batches/{batch_id}/participants/{participant_id}")
    
    # === ERROR HANDLING ===
    def test_add_participant_invalid_batch(self):
        """Test adding participant to non-existent batch returns 404"""
        fake_batch_id = "nonexistent_batch_12345"
        
        participant_data = {"fullName": "Test User"}
        
        response = self.session.post(
            f"{BASE_URL}/api/batches/{fake_batch_id}/participants",
            json=participant_data
        )
        assert response.status_code == 404, f"Expected 404, got {response.status_code}"
        print("✓ Add to invalid batch returns 404")
    
    def test_update_nonexistent_participant(self):
        """Test updating non-existent participant returns 404"""
        batch = self.get_or_create_test_batch()
        batch_id = batch["id"]
        fake_participant_id = "nonexistent_participant_12345"
        
        response = self.session.patch(
            f"{BASE_URL}/api/batches/{batch_id}/participants/{fake_participant_id}",
            json={"boarded": True}
        )
        assert response.status_code == 404, f"Expected 404, got {response.status_code}"
        print("✓ Update non-existent participant returns 404")


class TestParticipantCleanup:
    """Cleanup test data after tests"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Get auth token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        firebase_api_key = "AIzaSyCFl3n1PmSTN6eO4NdNDevIvFAmO78fV_M"
        auth_url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={firebase_api_key}"
        
        auth_response = self.session.post(auth_url, json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD,
            "returnSecureToken": True
        })
        
        if auth_response.status_code == 200:
            self.token = auth_response.json().get("idToken")
            self.session.headers.update({"Authorization": f"Bearer {self.token}"})
    
    def test_cleanup_test_participants(self):
        """Clean up TEST_ prefixed participants from all batches"""
        # Get all batches
        response = self.session.get(f"{BASE_URL}/api/batches")
        if response.status_code != 200:
            print("⚠ Could not get batches for cleanup")
            return
        
        batches = response.json()
        total_deleted = 0
        
        for batch in batches:
            batch_id = batch["id"]
            
            # Get participants
            part_response = self.session.get(f"{BASE_URL}/api/batches/{batch_id}/participants")
            if part_response.status_code != 200:
                continue
            
            participants = part_response.json()
            
            # Delete TEST_ prefixed participants
            for p in participants:
                if p.get("fullName", "").startswith("TEST_"):
                    delete_response = self.session.delete(
                        f"{BASE_URL}/api/batches/{batch_id}/participants/{p['id']}"
                    )
                    if delete_response.status_code == 200:
                        total_deleted += 1
        
        print(f"✓ Cleanup: Deleted {total_deleted} test participants")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
